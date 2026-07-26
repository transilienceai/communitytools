#!/usr/bin/env python3
"""Tests for tools/report_xlsx_build.py — the Transilience xlsx register builder.

Run: python3 tools/test_report_xlsx_build.py   (or via pytest)
Pure/offline: builds workbooks in-memory and asserts structure + values.
"""
import io
import unittest

import openpyxl

import report_xlsx_build as R


def _wb(data, **overrides):
    ov = {"target": None, "title": None, "location": None, "date": None}
    ov.update(overrides)
    wb, n = R.build_workbook(data, ov)
    # round-trip through bytes to prove it saves + reloads cleanly
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf), n


FULL = {
    "engagement": {
        "title": "External Penetration Test", "subtitle": "Consolidated Findings",
        "target": "ACME Corp", "sector": "BPO", "date": "2026-07-23",
        "report_id": "MERGE-20260723", "classification": "Confidential",
        "method": "PTES / OWASP / NIST SP 800-115",
    },
    "executive_summary": {
        "narrative": ["Para one.", "Para two."],
        "key_risks": ["risk a"],
        "positives": ["TLS 1.3 only", "MFA enforced"],
    },
    "coverage_table": {
        "header": ["Activity", "Description", "Status"],
        "rows": [["Port scan", "Full TCP 1-65535", "Completed"],
                 ["DNS testing", "Open recursion", "Completed"]],
        "note": "Rate-limited perimeter.",
    },
    "tools_used": [{"name": "nmap"}, {"name": "dig"}, "openssl"],
    "findings": [
        {"id": "F-001", "title": "Open DNS Recursion", "severity": "Medium",
         "cvss_score": 6.8, "cvss_vector": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:C/C:N/I:N/A:L",
         "cwe": "CWE-406", "owasp": "A05:2021",
         "affected": ["1.1.1.1:53/udp (Unbound)", "2.2.2.2:53/udp"],
         "description": "desc", "impact": "impact",
         "recommendation": "disable recursion",
         "references": ["US-CERT TA13-088A"],
         "poc": [{"description": "query", "command": "dig @1.1.1.1 x.com ANY"}],
         "needs_live_confirmation": False,
         "cves": [{"id": "CVE-2024-33655", "score": 7.5, "severity": "High"}]},
        {"id": "F-002", "title": "Proprietary Service", "severity": "Low",
         "cvss_score": 3.7, "affected": ["2.2.2.2: ~40 ports 63000-64900"],
         "needs_live_confirmation": True},
        {"id": "F-003", "title": "Patch Level Unverified", "severity": "Info",
         "cvss_score": None, "affected": ["1.1.1.1:53/udp", "2.2.2.2:53/udp"]},
    ],
}


class TestBuilder(unittest.TestCase):
    def test_sheets_and_roundtrip(self):
        wb, n = _wb(FULL)
        self.assertEqual(n, 3)
        self.assertEqual(wb.sheetnames,
                         ["Summary", "Findings Register", "Finding Detail",
                          "Host Inventory", "Coverage"])

    def test_summary_counts(self):
        wb, _ = _wb(FULL)
        ws = wb["Summary"]
        counts = {}
        for r in range(8, 13):
            counts[ws.cell(row=r, column=1).value] = ws.cell(row=r, column=2).value
        self.assertEqual(counts["Medium"], 1)
        self.assertEqual(counts["Low"], 1)
        self.assertEqual(counts["Informational"], 1)  # 'Info' -> 'Informational'
        self.assertEqual(counts["Critical"], 0)
        self.assertEqual(ws.cell(row=13, column=2).value, 3)  # total

    def test_register_rows_and_cvss(self):
        wb, _ = _wb(FULL)
        ws = wb["Findings Register"]
        self.assertEqual(ws.cell(row=8, column=1).value, "F-001")
        self.assertEqual(ws.cell(row=8, column=4).value, "6.8")     # cvss formatted
        self.assertEqual(ws.cell(row=10, column=4).value, "N/A")    # None -> N/A
        # CVE column populated for F-001
        self.assertIn("CVE-2024-33655", ws.cell(row=8, column=8).value)

    def test_host_inventory_derivation(self):
        wb, _ = _wb(FULL)
        ws = wb["Host Inventory"]
        # two distinct IPs derived from affected[]
        ips = [ws.cell(row=r, column=2).value for r in (8, 9)]
        self.assertEqual(set(ips), {"1.1.1.1", "2.2.2.2"})
        # 1.1.1.1 referenced by F-001 and F-003
        row_by_ip = {ws.cell(row=r, column=2).value: r for r in (8, 9)}
        f_col = ws.cell(row=row_by_ip["1.1.1.1"], column=4).value
        self.assertIn("F-001", f_col)
        self.assertIn("F-003", f_col)

    def test_status_from_needs_live(self):
        wb, _ = _wb(FULL)
        ws = wb["Finding Detail"]
        text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        self.assertIn("needs live confirmation", text)  # F-002
        self.assertIn("confirmed", text)                # F-001

    def test_coverage_table_and_positives(self):
        wb, _ = _wb(FULL)
        ws = wb["Coverage"]
        text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        self.assertIn("Port scan", text)
        self.assertIn("TLS 1.3 only", text)   # positive observation
        self.assertIn("nmap", text)           # tools_used

    def test_minimal_payload_no_crash(self):
        # only findings, no engagement/exec_summary/coverage — must still build
        wb, n = _wb({"findings": [{"id": "X-1", "title": "t", "severity": "High",
                                   "cvss_score": 8.1, "affected": ["9.9.9.9"]}]})
        self.assertEqual(n, 1)
        self.assertEqual(wb["Findings Register"].cell(row=8, column=3).value, "High")

    def test_empty_findings(self):
        wb, n = _wb({"engagement": {"title": "T"}, "findings": []})
        self.assertEqual(n, 0)
        self.assertEqual(wb["Summary"].cell(row=13, column=2).value, 0)

    def test_override_client(self):
        wb, _ = _wb({"engagement": {"title": "T", "target": "Old"},
                     "findings": []}, target="NewClient")
        ctx = wb["Summary"].cell(row=3, column=1).value
        self.assertIn("NewClient", ctx)


if __name__ == "__main__":
    unittest.main(verbosity=2)
