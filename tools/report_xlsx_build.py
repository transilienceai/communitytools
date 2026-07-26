#!/usr/bin/env python3
"""Deterministic Transilience-style XLSX findings-register builder.

Reads ONE report_data.json (the same canonical payload
skills/transilience-report-style/reference/generate_report.py renders to PDF —
see report-data-schema.json) and emits a multi-sheet Excel workbook in
Transilience house style: Summary, Findings Register, Finding Detail,
Host Inventory, Coverage.

No LLM, no network, stdlib + openpyxl only, clockless except an optional
--date override (falls back to engagement.date). Given the same report_data.json
it always produces the same workbook, so it slots into the merge-reports /
pentest-engagement finalize chain right after generate_report.py.

REUSABLE CONTENT: this tool carries NO client name, host, or engagement path.
Every client-specific value is read from the report_data.json passed in.

Usage:
    python3 tools/report_xlsx_build.py <report_data.json> -o <out.xlsx>
                                       [--client NAME] [--title TITLE]
                                       [--location LOC] [--date DATE]
"""
import argparse
import json
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    raise

# ---- house palette (extracted from the Transilience xlsx house style) -------
BRAND = "7A1E46"          # Transilience maroon — headers, title, section bars
SUBTITLE = "1F3864"       # dark blue — subtitle line
LABEL_FILL = "F5F2F7"     # light lavender — detail label cells
BORDER = "C7D0DA"         # thin grid line
SEV_FILL = {
    "Critical": "7B0000",
    "High": "C00000",
    "Medium": "E08600",
    "Low": "C9A800",
    "Info": "3F8F3F",
    "Informational": "3F8F3F",
}
SEV_ORDER = ["Critical", "High", "Medium", "Low", "Informational"]
WHITE = "FFFFFF"

_thin = Side(style="thin", color=BORDER)
GRID = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def norm_sev(s):
    """Canonicalize a severity label; 'Info' displays as 'Informational'."""
    s = (s or "").strip().title()
    if s in ("Info", "Informational"):
        return "Informational"
    return s if s in ("Critical", "High", "Medium", "Low") else "Informational"


def sev_display(s):
    return norm_sev(s)


# ---- low-level cell helpers -------------------------------------------------
def _c(ws, r, col, value, *, bold=False, size=11, color=None, fill=None,
       align=None, wrap=False, border=False, name=None):
    cell = ws.cell(row=r, column=col, value=value)
    cell.font = Font(name=name, bold=bold, size=size, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(
        horizontal=align, vertical="top", wrap_text=wrap)
    if border:
        cell.border = GRID
    return cell


def _brand_header(ws, subtitle, ctx3, ctx4, span=6):
    """Rows 1-4: TRANSILIENCE / subtitle / two context lines, merged across."""
    _c(ws, 1, 1, "TRANSILIENCE", bold=True, size=16, color=BRAND)
    _c(ws, 2, 1, subtitle, bold=True, size=11, color=SUBTITLE)
    _c(ws, 3, 1, ctx3, size=10, color="404040")
    _c(ws, 4, 1, ctx4, size=9, color="808080")
    for r in (1, 2, 3, 4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)


def _table_header(ws, r, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        _c(ws, r, i, h, bold=True, size=10, color=WHITE, fill=BRAND,
           align="center", wrap=True, border=True)
        ws.column_dimensions[get_column_letter(i)].width = w


def _sev_cell(ws, r, col, sev):
    disp = sev_display(sev)
    _c(ws, r, col, disp, bold=True, size=10, color=WHITE,
       fill=SEV_FILL.get(disp, SEV_FILL["Informational"]),
       align="center", wrap=True, border=True)


# ---- report_data helpers ----------------------------------------------------
_IP_RE = re.compile(r"(\d{1,3}(?:\.\d{1,3}){3})")


def cvss_str(f):
    v = f.get("cvss_score")
    if v is None or v == "" or (isinstance(v, str) and v.strip().lower() in ("n/a", "none")):
        return "N/A"
    try:
        return f"{float(v):.1f}"
    except (TypeError, ValueError):
        return str(v)


def cvss_version(f):
    """CVSS version tag ('v4.0'/'v3.1'/'v3.0'/'v2.0') from the finding's vector
    prefix, or '' when unknown. v4.0 is the primary version repo-wide; a
    prefix-less vector is CVSS v2.0."""
    vec = f.get("cvss_vector") or ""
    m = re.match(r"(?i)^CVSS:(\d\.\d)/", str(vec))
    if m:
        return "v" + m.group(1)
    return "v2.0" if vec else ""


def cvss_label(f):
    """'CVSS v4.0 9.3' — score paired with its version when both are known."""
    score = cvss_str(f)
    ver = cvss_version(f)
    if score == "N/A":
        return "CVSS N/A"
    return f"CVSS {ver} {score}".replace("  ", " ")


def classification(f):
    bits = [b for b in (f.get("cwe"), f.get("owasp")) if b]
    return "; ".join(bits) if bits else "—"


def status_of(f):
    return "Open — needs live confirmation" if f.get("needs_live_confirmation") else "Open — confirmed"


def refs_str(f):
    refs = f.get("references") or []
    return "; ".join(str(x) for x in refs) if refs else "—"


def cves_str(f):
    out = []
    for c in (f.get("cves") or []):
        cid = c.get("id") if isinstance(c, dict) else c
        if cid:
            out.append(str(cid))
    return ", ".join(out)


def poc_text(f):
    """Render the ordered PoC steps as a single readable cell."""
    lines = []
    for i, step in enumerate(f.get("poc") or [], start=1):
        if not isinstance(step, dict):
            continue
        desc = (step.get("description") or "").strip()
        cmd = (step.get("command") or "").strip()
        seg = f"{i}. {desc}" if desc else f"{i}."
        if cmd:
            seg += f"\n    $ {cmd}"
        lines.append(seg)
    return "\n".join(lines) if lines else "See evidence appendix / proof package."


def affected_list(f):
    return [str(a) for a in (f.get("affected") or []) if str(a).strip()]


# ---- sheet builders ---------------------------------------------------------
def build_summary(wb, eng, findings, ctx3, ctx4):
    ws = wb.create_sheet("Summary")
    for col, w in zip("ABCDEF", (24, 16, 16, 16, 16, 16)):
        ws.column_dimensions[col].width = w
    _brand_header(ws, f"{eng.get('subtitle') or 'Security Assessment'} — Executive Summary", ctx3, ctx4)

    counts = {s: 0 for s in SEV_ORDER}
    for f in findings:
        counts[norm_sev(f.get("severity"))] += 1
    r = 7
    _c(ws, r, 1, "Severity", bold=True, size=10, color=WHITE, fill=BRAND, align="center", wrap=True, border=True)
    _c(ws, r, 2, "Count", bold=True, size=10, color=WHITE, fill=BRAND, align="center", wrap=True, border=True)
    r += 1
    for s in SEV_ORDER:
        _c(ws, r, 1, s, bold=True, size=10, color=WHITE, fill=SEV_FILL[s], align="left", border=True)
        _c(ws, r, 2, counts[s], align="center", border=True)
        r += 1
    _c(ws, r, 1, "Total", bold=True, border=True)
    _c(ws, r, 2, len(findings), bold=True, align="center", border=True)

    r += 2
    _c(ws, r, 1, "Scope & Method", bold=True, size=11, color=BRAND)
    r += 1
    scope_rows = [
        ("Client / Target", eng.get("target") or "—"),
        ("Sector", eng.get("sector") or "—"),
        ("Method", eng.get("method") or "—"),
        ("Report ID", eng.get("report_id") or "—"),
        ("Date", eng.get("date") or "—"),
        ("Classification", eng.get("classification") or eng.get("confidentiality") or "Confidential"),
    ]
    for k, v in scope_rows:
        _c(ws, r, 1, k, bold=True, size=10, color=BRAND, fill=LABEL_FILL, border=True)
        _c(ws, r, 2, str(v), wrap=True, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1
    return ws


def build_register(wb, findings, ctx3, ctx4):
    ws = wb.create_sheet("Findings Register")
    _brand_header(ws, "Findings Register", ctx3, ctx4)
    headers = ["ID", "Finding", "Severity", "CVSS", "CVSS Vector",
               "Affected Hosts", "CWE / Classification", "CVE(s)", "Status", "References"]
    widths = [8, 40, 14, 7, 32, 26, 30, 20, 24, 34]
    _table_header(ws, 7, headers, widths)
    ws.freeze_panes = "A8"
    r = 8
    for f in findings:
        _c(ws, r, 1, f.get("id") or "", bold=True, align="center", border=True)
        _c(ws, r, 2, f.get("title") or "", wrap=True, border=True)
        _sev_cell(ws, r, 3, f.get("severity"))
        _c(ws, r, 4, cvss_str(f), align="center", border=True)
        _c(ws, r, 5, f.get("cvss_vector") or "—", wrap=True, border=True)
        _c(ws, r, 6, "\n".join(affected_list(f)) or "—", wrap=True, border=True)
        _c(ws, r, 7, classification(f), wrap=True, border=True)
        _c(ws, r, 8, cves_str(f) or "—", wrap=True, border=True)
        _c(ws, r, 9, status_of(f), wrap=True, border=True)
        _c(ws, r, 10, refs_str(f), wrap=True, border=True)
        r += 1
    return ws


def build_detail(wb, findings, ctx3, ctx4):
    ws = wb.create_sheet("Finding Detail")
    for col, w in zip("ABCDEF", (22, 20, 18, 18, 18, 18)):
        ws.column_dimensions[col].width = w
    _brand_header(ws, "Findings — Full Detail", ctx3, ctx4)
    r = 6
    for f in findings:
        r += 1
        title = f"{f.get('id') or ''} — {f.get('title') or ''}".strip(" —")
        _c(ws, r, 1, title, bold=True, size=12, color=WHITE, fill=BRAND, align="left", wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        rows = [
            ("Severity", f"{sev_display(f.get('severity'))} ({cvss_label(f)})"),
            ("CVSS Vector", f.get("cvss_vector") or "N/A"),
            ("Affected Hosts", "\n".join(affected_list(f)) or "—"),
            ("Classification", classification(f)),
            ("CVE(s)", cves_str(f) or "—"),
            ("Status", status_of(f)),
            ("Description", f.get("description") or "—"),
            ("Impact", f.get("impact") or "—"),
            ("Evidence / PoC", poc_text(f)),
            ("Recommendation", f.get("recommendation") or "—"),
            ("References", refs_str(f)),
        ]
        if f.get("calibration"):
            rows.append(("Assessor Note", f.get("calibration")))
        if f.get("cve_caveat"):
            rows.append(("CVE Caveat", f.get("cve_caveat")))
        for k, v in rows:
            _c(ws, r, 1, k, bold=True, size=10, color=BRAND, fill=LABEL_FILL, wrap=True, border=True)
            _c(ws, r, 2, str(v), wrap=True, border=True)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            r += 1
        r += 1  # blank spacer between findings
    return ws


def build_host_inventory(wb, findings, ctx3, ctx4):
    """Derive a host/port inventory from every finding's affected[] entries."""
    ws = wb.create_sheet("Host Inventory")
    _brand_header(ws, "Host & Port Inventory", ctx3, ctx4)
    headers = ["#", "IP Address", "Ports / Services (as reported)", "Findings"]
    widths = [5, 18, 60, 22]
    _table_header(ws, 7, headers, widths)
    ws.freeze_panes = "A8"

    hosts = {}   # ip -> {"notes": set, "findings": set}
    order = []
    for f in findings:
        fid = f.get("id") or ""
        for a in affected_list(f):
            m = _IP_RE.search(a)
            ip = m.group(1) if m else a.split()[0]
            if ip not in hosts:
                hosts[ip] = {"notes": [], "findings": set()}
                order.append(ip)
            # the affected string minus the leading IP is the port/service note
            note = a
            if m:
                note = a.replace(m.group(1), "", 1).lstrip(" :,-")
            note = note.strip() or "(see finding)"
            if note not in hosts[ip]["notes"]:
                hosts[ip]["notes"].append(note)
            if fid:
                hosts[ip]["findings"].add(fid)

    r = 8
    for i, ip in enumerate(order, start=1):
        _c(ws, r, 1, i, align="center", border=True)
        _c(ws, r, 2, ip, border=True)
        _c(ws, r, 3, "\n".join(hosts[ip]["notes"]) or "—", wrap=True, border=True)
        _c(ws, r, 4, ", ".join(sorted(hosts[ip]["findings"])) or "—", wrap=True, border=True)
        r += 1
    return ws


def build_coverage(wb, data, eng, ctx3, ctx4):
    ws = wb.create_sheet("Coverage")
    for col, w in zip("ABC", (34, 74, 22)):
        ws.column_dimensions[col].width = w
    _brand_header(ws, "Methodology, Coverage & Assurance", ctx3, ctx4, span=3)
    r = 7
    cov = data.get("coverage_table") or {}
    header = cov.get("header") or ["Activity", "Description", "Status"]
    _table_header(ws, r, header[:3], [34, 74, 22][:len(header[:3])])
    r += 1
    rows = cov.get("rows") or []
    if not rows:
        # No structured coverage table in the payload — record the method only,
        # honestly, rather than fabricating activity claims.
        rows = [[eng.get("method") or "Assessment methodology", "See report body", "Completed"]]
    for row in rows:
        cells = list(row) + [""] * (3 - len(row))
        for ci in range(3):
            _c(ws, r, ci + 1, str(cells[ci]), wrap=True, border=True,
               align="center" if ci == 2 else None)
        r += 1
    if cov.get("note"):
        r += 1
        _c(ws, r, 1, str(cov["note"]), size=9, color="808080", wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    # Assurance / positive observations
    positives = (data.get("executive_summary") or {}).get("positives") or []
    if positives:
        r += 1
        _c(ws, r, 1, "Assurance / Positive Observations", bold=True, size=11, color=BRAND)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
        for p in positives:
            _c(ws, r, 1, f"• {p}", wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1

    tools = data.get("tools_used") or []
    if tools:
        r += 1
        names = [t.get("name") if isinstance(t, dict) else str(t) for t in tools]
        _c(ws, r, 1, "Tools used", bold=True, size=10, color=BRAND, fill=LABEL_FILL, border=True)
        _c(ws, r, 2, ", ".join(n for n in names if n), wrap=True, border=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    return ws


def build_workbook(data, overrides):
    eng = dict(data.get("engagement") or {})
    for k in ("title", "target", "date"):
        if overrides.get(k):
            eng[k] = overrides[k]
    findings = list(data.get("findings") or [])

    client = eng.get("target") or ""
    loc = overrides.get("location") or eng.get("sector") or ""
    title = eng.get("title") or "Security Assessment"
    ctx3 = "  |  ".join([x for x in (client, loc, title) if x]) or title
    ctx4_bits = []
    if eng.get("report_id"):
        ctx4_bits.append(f"Ref {eng['report_id']}")
    if eng.get("date"):
        ctx4_bits.append(f"Tested {eng['date']}")
    ctx4_bits.append((eng.get("classification") or eng.get("confidentiality") or "Confidential").upper())
    ctx4 = "  |  ".join(ctx4_bits)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default sheet
    build_summary(wb, eng, findings, ctx3, ctx4)
    build_register(wb, findings, ctx3, ctx4)
    build_detail(wb, findings, ctx3, ctx4)
    build_host_inventory(wb, findings, ctx3, ctx4)
    build_coverage(wb, data, eng, ctx3, ctx4)
    return wb, len(findings)


def main():
    ap = argparse.ArgumentParser(description="Build a Transilience-style XLSX from a report_data.json")
    ap.add_argument("report_data", help="path to report_data.json")
    ap.add_argument("-o", "--output", required=True, help="output .xlsx path")
    ap.add_argument("--client", help="override engagement.target (client name)")
    ap.add_argument("--title", help="override engagement.title")
    ap.add_argument("--location", help="location shown in the context line")
    ap.add_argument("--date", help="override engagement.date")
    args = ap.parse_args()

    with open(args.report_data, encoding="utf-8") as fh:
        data = json.load(fh)
    overrides = {"target": args.client, "title": args.title,
                 "location": args.location, "date": args.date}
    wb, n = build_workbook(data, overrides)
    wb.save(args.output)
    print(json.dumps({"ok": True, "out": args.output, "findings": n,
                      "sheets": wb.sheetnames}))


if __name__ == "__main__":
    main()
